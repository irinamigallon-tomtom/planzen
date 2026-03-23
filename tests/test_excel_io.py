"""
Tests for excel_io.py — values output and formula output.

Includes parametrized tests that vary the number of epics to ensure
formula row/column references are dynamically correct regardless of
sheet shape.
"""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
import pytest
from openpyxl.utils import get_column_letter

from planzen.config import (
    FISCAL_QUARTERS,
    OUT_COL_BUDGET_BUCKET,
    LABEL_CAPACITY_ALERT_ROW,
    LABEL_ENG_ABSENCE,
    LABEL_ENG_BRUTO,
    LABEL_ENG_NET,
    LABEL_MGMT_ABSENCE,
    LABEL_MGMT_CAPACITY,
    LABEL_MGMT_NET,
    LABEL_TOTAL_ROW,
    OUT_COL_EPIC,
    OUT_COL_ESTIMATION,
    OUT_COL_OFF_ESTIMATE,
    OUT_COL_TOTAL_WEEKS,
)
from planzen.core_logic import CapacityConfig, build_output_table
from planzen.excel_io import formulas_path, read_input, validate_input_file, write_output, write_output_with_formulas

# ---------------------------------------------------------------------------
# Shared fixtures and helpers
# ---------------------------------------------------------------------------

_CAPACITY = CapacityConfig(num_engineers=5, num_managers=2)
_Q1_START = FISCAL_QUARTERS[1][0]
_START = _Q1_START
_END = _Q1_START + timedelta(weeks=3)  # 4 Mondays

_EPICS = pd.DataFrame([
    {"Epic Description": "Epic A", "Estimation": 10.0, "Budget Bucket": "Core",
     "Type": "Feature", "Link": "https://jira.example.com/A", "Priority": 0, "Milestone": "Q1"},
    {"Epic Description": "Epic B", "Estimation": 5.0,  "Budget Bucket": "Core",
     "Type": "Feature", "Link": "https://jira.example.com/B", "Priority": 1, "Milestone": "Q1"},
])

_CAPACITY_LABELS = {
    LABEL_ENG_BRUTO, LABEL_ENG_ABSENCE, LABEL_ENG_NET,
    LABEL_MGMT_CAPACITY, LABEL_MGMT_ABSENCE, LABEL_MGMT_NET,
    LABEL_TOTAL_ROW, LABEL_CAPACITY_ALERT_ROW,
}


def _make_epics(n: int) -> pd.DataFrame:
    """Build a DataFrame with *n* epics of varying estimations."""
    return pd.DataFrame([
        {
            "Epic Description": f"Epic {i}",
            "Estimation": float(10 * (i + 1)),
            "Budget Bucket": "Core",
            "Type": "Feature",
            "Link": f"https://jira.example.com/{i}",
            "Priority": i,
            "Milestone": "Q1",
        }
        for i in range(n)
    ])


@pytest.fixture()
def output_df() -> pd.DataFrame:
    return build_output_table(_EPICS, _CAPACITY, _START, _END)


@pytest.fixture()
def values_file(tmp_path: Path, output_df: pd.DataFrame) -> Path:
    p = tmp_path / "output.xlsx"
    write_output(output_df, p)
    return p


@pytest.fixture()
def formulas_file(tmp_path: Path, output_df: pd.DataFrame) -> Path:
    p = tmp_path / "output_formulas.xlsx"
    n_base = len([c for c in output_df.columns if c not in {"Budget Bucket", "Epic Description", "Priority", "Estimation", "Total Weeks", "Off Estimate"}])
    write_output_with_formulas(output_df, p, n_base)
    return p


def _find_row(ws, label: str, label_col: int = 2) -> int:
    """Return the 1-based Excel row number whose label_col cell equals *label*."""
    for row in ws.iter_rows():
        if row[label_col - 1].value == label:
            return row[0].row
    raise KeyError(f"Label {label!r} not found in worksheet")


def _week_cols(ws, header_row: int = 1) -> list[int]:
    """Return 1-based column indices for all week columns (non-metadata columns)."""
    non_week = {"Budget Bucket", "Epic Description", "Priority", "Estimation", "Total Weeks", "Off Estimate"}
    return [
        cell.column for cell in ws[header_row]
        if cell.value not in non_week and cell.value is not None
    ]


def _total_weeks_col(ws, header_row: int = 1) -> int:
    for cell in ws[header_row]:
        if cell.value == "Total Weeks":
            return cell.column
    raise KeyError("Total Weeks column not found")


# ---------------------------------------------------------------------------
# formulas_path helper
# ---------------------------------------------------------------------------

def test_formulas_path_appends_suffix() -> None:
    assert formulas_path(Path("out/result.xlsx")) == Path("out/result_formulas.xlsx")


def test_formulas_path_preserves_directory() -> None:
    p = formulas_path(Path("/data/examples/output_example.xlsx"))
    assert p == Path("/data/examples/output_example_formulas.xlsx")


# ---------------------------------------------------------------------------
# read_plan: column validation and flexibility
# ---------------------------------------------------------------------------

def _base_row(**extra) -> dict:
    return {
        "Epic Description": "E", "Estimation": 1.0, "Budget Bucket": "B",
        "Type": "Feature", "Link": "http://x", "Priority": 0,
        **extra,
    }


def _config_rows() -> list[dict]:
    return [
        {"Budget Bucket": "Engineer Capacity (Bruto)", "Estimation": 5.0},
        {"Budget Bucket": "Management Capacity (Bruto)", "Estimation": 2.0},
    ]


def _write_input(path: Path, epics: list[dict], config: list[dict] | None = None) -> None:
    rows = (config if config is not None else _config_rows()) + epics
    pd.DataFrame(rows).to_excel(path, index=False)


# ---------------------------------------------------------------------------
# read_input: team config + epic parsing
# ---------------------------------------------------------------------------

def test_read_input_returns_engineers_and_managers(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    _write_input(p, [_base_row()])
    _, cap = read_input(p, 2)
    eng, mgr = cap.num_engineers, cap.num_managers
    assert eng == 5.0
    assert mgr == 2.0


def test_read_input_returns_none_when_absence_omitted(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    _write_input(p, [_base_row()])
    _, cap = read_input(p, 2)
    eng_abs, mgmt_abs = cap.eng_absence_per_week, cap.mgmt_absence_per_week
    assert eng_abs is None
    assert mgmt_abs is None


def test_read_input_parses_optional_absence_days(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    config = _config_rows() + [
        {"Budget Bucket": "Engineer Absence", "Estimation": 10.0},
        {"Budget Bucket": "Management Absence", "Estimation": 4.0},
    ]
    _write_input(p, [_base_row()], config=config)
    _, cap = read_input(p, 2)
    # 10 days / 5 days-per-week / 13 Q2-weeks
    assert cap.eng_absence_per_week == pytest.approx(10.0 / 5 / 13)
    assert cap.mgmt_absence_per_week == pytest.approx(4.0 / 5 / 13)


def test_read_input_epic_rows_exclude_config_rows(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    _write_input(p, [_base_row()])
    epics, _ = read_input(p, 2)
    assert len(epics) == 1


def test_read_input_accepts_extra_columns(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    _write_input(p, [_base_row(**{"Extra Col": "ignored"})])
    epics, _ = read_input(p, 2)
    assert "Extra Col" in epics.columns


def test_read_input_accepts_optional_milestone(tmp_path: Path) -> None:
    p_with = tmp_path / "with.xlsx"
    p_without = tmp_path / "without.xlsx"
    _write_input(p_with, [_base_row(Milestone="Q1")])
    _write_input(p_without, [_base_row()])
    read_input(p_with, 2)
    read_input(p_without, 2)


def test_read_input_preserves_column_order(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    # Config rows come first; their keys (Epic Description, Estimation) set the
    # initial column order.  Epic-only keys follow in insertion order.
    rows = _config_rows() + [{
        "Epic Description": "E", "Estimation": 1.0,
        "Priority": 0, "Link": "http://x",
        "Type": "Feature", "Budget Bucket": "B",
    }]
    pd.DataFrame(rows).to_excel(p, index=False)
    epics, _ = read_input(p, 2)
    # Column order must match the file: Budget Bucket and Estimation come first
    # (from config rows), then the remaining epic columns.
    assert epics.columns[0] == "Budget Bucket"
    assert "Epic Description" in epics.columns
    # All required epic columns must be present (Type is optional)
    for col in ("Priority", "Link", "Budget Bucket"):
        assert col in epics.columns


def test_read_input_raises_for_missing_engineers_row(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    config = [{"Budget Bucket": "Management Capacity (Bruto)", "Estimation": 2.0}]
    _write_input(p, [_base_row()], config=config)
    with pytest.raises(ValueError, match="Engineer Capacity"):
        read_input(p, 2)


def test_read_input_raises_for_missing_required_epic_columns(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    # Config rows identified only by Epic Description; no Budget Bucket column at all.
    config = [{"Epic Description": "Engineer Capacity (Bruto)", "Estimation": 5.0}]
    rows = config + [{"Epic Description": "E", "Estimation": 1.0}]
    pd.DataFrame(rows).to_excel(p, index=False)
    with pytest.raises(ValueError, match="missing required columns"):
        read_input(p, 2)


# ---------------------------------------------------------------------------
# read_input: fuzzy / case-insensitive config label matching
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("eng_label", [
    "Engineer Capacity (Bruto)",     # canonical
    "engineer capacity (bruto)",     # all lowercase
    "ENGINEER CAPACITY (BRUTO)",     # all uppercase
    "  Engineer Capacity (Bruto) ",  # extra whitespace
    "Engineer Capacity",             # without parenthetical — still fuzzy-matches
])
def test_read_input_accepts_engineer_label_variants(tmp_path: Path, eng_label: str) -> None:
    p = tmp_path / "input.xlsx"
    config = [
        {"Budget Bucket": eng_label, "Estimation": 4.0},
        {"Budget Bucket": "Management Capacity (Bruto)", "Estimation": 2.0},
    ]
    _write_input(p, [_base_row()], config=config)
    _, cap = read_input(p, 2)
    eng = cap.num_engineers
    assert eng == 4.0


@pytest.mark.parametrize("mgr_label", [
    "Management Capacity (Bruto)",   # canonical
    "management capacity (bruto)",   # lowercase
    "Management Capacity",           # without parenthetical
])
def test_read_input_accepts_manager_label_variants(tmp_path: Path, mgr_label: str) -> None:
    p = tmp_path / "input.xlsx"
    config = [
        {"Budget Bucket": "Engineer Capacity (Bruto)", "Estimation": 5.0},
        {"Budget Bucket": mgr_label, "Estimation": 2.0},
    ]
    _write_input(p, [_base_row()], config=config)
    _, cap = read_input(p, 2)
    mgr = cap.num_managers
    assert mgr == 2.0


@pytest.mark.parametrize("eng_abs_label,mgr_abs_label", [
    ("Engineer Absence", "Management Absence"),               # canonical
    ("Engineer Absence (days)", "Management Absence (days)"), # with (days) — fuzzy strips it
    ("engineer absence", "management absence"),               # lowercase
])
def test_read_input_accepts_absence_label_variants(
    tmp_path: Path, eng_abs_label: str, mgr_abs_label: str
) -> None:
    p = tmp_path / "input.xlsx"
    config = _config_rows() + [
        {"Budget Bucket": eng_abs_label, "Estimation": 10.0},
        {"Budget Bucket": mgr_abs_label, "Estimation": 4.0},
    ]
    _write_input(p, [_base_row()], config=config)
    _, cap = read_input(p, 2)
    assert cap.eng_absence_per_week == pytest.approx(10.0 / 5 / 13)
    assert cap.mgmt_absence_per_week == pytest.approx(4.0 / 5 / 13)



# ---------------------------------------------------------------------------
# read_input: empty row and unnamed row handling
# ---------------------------------------------------------------------------

def test_read_input_discards_fully_empty_rows(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    rows = _config_rows() + [
        _base_row(),
        # fully empty row
        {"Epic Description": None, "Estimation": None, "Budget Bucket": None,
         "Type": None, "Link": None, "Priority": None},
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    epics, _ = read_input(p, 2)
    assert len(epics) == 1
    assert epics.iloc[0]["Epic Description"] == "E"


def test_read_input_discards_rows_without_epic_description(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    rows = _config_rows() + [
        _base_row(),
        # row with data but no Epic Description
        {"Epic Description": None, "Estimation": 3.0, "Budget Bucket": "Core",
         "Type": "Feature", "Link": "http://x", "Priority": 2},
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    epics, _ = read_input(p, 2)
    assert len(epics) == 1


def test_read_input_logs_warning_for_unnamed_rows(tmp_path: Path, caplog) -> None:
    import logging
    p = tmp_path / "input.xlsx"
    rows = _config_rows() + [
        _base_row(),
        {"Epic Description": None, "Estimation": 3.0, "Budget Bucket": "Core",
         "Type": "Feature", "Link": "http://x", "Priority": 2},
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    with caplog.at_level(logging.WARNING, logger="planzen.excel_io"):
        read_input(p, 2)
    assert any("Discarding" in r.message for r in caplog.records)


def test_read_input_handles_mixed_empty_and_valid_rows(tmp_path: Path) -> None:
    """Empty rows scattered between valid rows are all silently dropped."""
    p = tmp_path / "input.xlsx"
    rows = _config_rows() + [
        _base_row(**{"Epic Description": "A", "Priority": 1}),
        {"Epic Description": None, "Estimation": None},  # empty
        _base_row(**{"Epic Description": "B", "Priority": 2}),
        {"Epic Description": None, "Estimation": None},  # empty
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    epics, _ = read_input(p, 2)
    assert list(epics["Epic Description"]) == ["A", "B"]


# ---------------------------------------------------------------------------
# validate_input_file: user-friendly error collection
# ---------------------------------------------------------------------------

def test_validate_returns_empty_for_valid_file(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    _write_input(p, [_base_row()])
    assert validate_input_file(p) == []


def test_validate_reports_unreadable_file(tmp_path: Path) -> None:
    p = tmp_path / "missing.xlsx"
    errors = validate_input_file(p)
    assert len(errors) == 1
    assert "not found" in errors[0].lower() or "cannot" in errors[0].lower()


def test_validate_reports_missing_epic_description_column(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    pd.DataFrame([{"Estimation": 1.0}]).to_excel(p, index=False)
    errors = validate_input_file(p)
    assert any("Epic Description" in e for e in errors)


def test_validate_reports_missing_engineers_config_row(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    config = [{"Budget Bucket": "Management Capacity (Bruto)", "Estimation": 2.0}]
    _write_input(p, [_base_row()], config=config)
    errors = validate_input_file(p)
    assert any("Engineer Capacity" in e for e in errors)


def test_validate_reports_missing_managers_config_row(tmp_path: Path) -> None:
    """Management capacity is optional — no error when absent; defaults to 1.0."""
    p = tmp_path / "input.xlsx"
    config = [{"Budget Bucket": "Engineer Capacity (Bruto)", "Estimation": 5.0}]
    _write_input(p, [_base_row()], config=config)
    errors = validate_input_file(p)
    assert not any("Management Capacity" in e for e in errors)


def test_validate_reports_non_positive_engineers(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    config = [
        {"Budget Bucket": "Engineer Capacity (Bruto)", "Estimation": 0},
        {"Budget Bucket": "Management Capacity (Bruto)", "Estimation": 2.0},
    ]
    _write_input(p, [_base_row()], config=config)
    errors = validate_input_file(p)
    assert any("Engineer Capacity" in e for e in errors)
    assert any("greater than 0" in e or "positive" in e.lower() for e in errors)


def test_validate_reports_non_numeric_engineers(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    config = [
        {"Budget Bucket": "Engineer Capacity (Bruto)", "Estimation": "lots"},
        {"Budget Bucket": "Management Capacity (Bruto)", "Estimation": 2.0},
    ]
    _write_input(p, [_base_row()], config=config)
    errors = validate_input_file(p)
    assert any("Engineer Capacity" in e for e in errors)


def test_validate_reports_negative_absence_days(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    config = _config_rows() + [
        {"Budget Bucket": "Engineer Absence", "Estimation": -3},
    ]
    _write_input(p, [_base_row()], config=config)
    errors = validate_input_file(p)
    assert any("Absence" in e for e in errors)
    assert any("negative" in e.lower() or ">= 0" in e for e in errors)


def test_validate_reports_missing_required_epic_columns(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    # Budget Bucket column absent → must be reported as missing required column.
    config = [{"Epic Description": "Engineer Capacity (Bruto)", "Estimation": 5.0}]
    rows = config + [{"Epic Description": "E", "Estimation": 1.0}]
    pd.DataFrame(rows).to_excel(p, index=False)
    errors = validate_input_file(p)
    assert any("Budget Bucket" in e for e in errors)


def test_validate_reports_no_epic_rows(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    _write_input(p, [], config=_config_rows())
    errors = validate_input_file(p)
    assert any("epic" in e.lower() for e in errors)


def test_validate_reports_non_numeric_estimation(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    _write_input(p, [_base_row(**{"Estimation": "TBD"})])
    errors = validate_input_file(p)
    assert any("Estimation" in e for e in errors)
    assert any("numeric" in e.lower() or "number" in e.lower() for e in errors)


def test_validate_reports_negative_estimation(tmp_path: Path) -> None:
    p = tmp_path / "input.xlsx"
    _write_input(p, [_base_row(**{"Estimation": -1.0})])
    errors = validate_input_file(p)
    assert any("Estimation" in e for e in errors)


def test_validate_no_error_for_blank_priority(tmp_path: Path) -> None:
    """Priority is imputed from Budget Bucket; a blank value is not an error."""
    p = tmp_path / "input.xlsx"
    row = _base_row(**{"Budget Bucket": "Customer Support"})
    row["Priority"] = None
    _write_input(p, [row])
    errors = validate_input_file(p)
    assert not any("Priority" in e for e in errors)


def test_validate_collects_multiple_errors(tmp_path: Path) -> None:
    """All issues are reported at once, not just the first one."""
    p = tmp_path / "input.xlsx"
    # No config rows at all, and epic has bad Estimation
    rows = [{"Epic Description": "E", "Estimation": "TBD", "Budget Bucket": "B",
             "Type": "F", "Link": "http://x", "Priority": 0}]
    pd.DataFrame(rows).to_excel(p, index=False)
    errors = validate_input_file(p)
    assert len(errors) >= 2  # at least: missing engineers row + bad estimation


# ---------------------------------------------------------------------------
# Values file: no formula strings
# ---------------------------------------------------------------------------

def test_values_file_is_created(values_file: Path) -> None:
    assert values_file.exists()


def test_values_file_net_capacity_cells_are_numeric(values_file: Path) -> None:
    wb = openpyxl.load_workbook(values_file, data_only=True)
    ws = wb.active
    eng_net_row = _find_row(ws, LABEL_ENG_NET)
    mgmt_net_row = _find_row(ws, LABEL_MGMT_NET)
    for col_idx in _week_cols(ws):
        assert isinstance(ws.cell(eng_net_row, col_idx).value, (int, float)), (
            f"Eng Net cell ({eng_net_row},{col_idx}) should be numeric in values file"
        )
        assert isinstance(ws.cell(mgmt_net_row, col_idx).value, (int, float)), (
            f"Mgmt Net cell ({mgmt_net_row},{col_idx}) should be numeric in values file"
        )


def test_values_file_total_weeks_is_numeric(values_file: Path, output_df: pd.DataFrame) -> None:
    wb = openpyxl.load_workbook(values_file, data_only=True)
    ws = wb.active
    tw_col = _total_weeks_col(ws)
    for row in ws.iter_rows(min_row=2):
        label = row[1].value  # column B
        if label not in _CAPACITY_LABELS and label is not None:
            cell_val = ws.cell(row[0].row, tw_col).value
            assert isinstance(cell_val, (int, float)), (
                f"Total Weeks for '{label}' should be numeric in values file, got {cell_val!r}"
            )


def test_values_file_weekly_allocation_is_numeric(values_file: Path) -> None:
    wb = openpyxl.load_workbook(values_file, data_only=True)
    ws = wb.active
    total_row = _find_row(ws, LABEL_TOTAL_ROW)
    for col_idx in _week_cols(ws):
        val = ws.cell(total_row, col_idx).value
        assert isinstance(val, (int, float)), (
            f"Weekly Allocation cell ({total_row},{col_idx}) should be numeric, got {val!r}"
        )


# ---------------------------------------------------------------------------
# Formula file: specific cells contain Excel formulas
# ---------------------------------------------------------------------------

def test_formulas_file_is_created(formulas_file: Path) -> None:
    assert formulas_file.exists()


def test_formulas_file_eng_net_has_subtraction_formula(formulas_file: Path) -> None:
    wb = openpyxl.load_workbook(formulas_file, data_only=False)
    ws = wb.active
    eng_net_row = _find_row(ws, LABEL_ENG_NET)
    for col_idx in _week_cols(ws):
        val = ws.cell(eng_net_row, col_idx).value
        assert isinstance(val, str) and val.startswith("="), (
            f"Eng Net cell ({eng_net_row},{col_idx}) should be a formula, got {val!r}"
        )
        assert "-" in val, f"Eng Net formula should subtract absence: {val!r}"


def test_formulas_file_mgmt_net_has_subtraction_formula(formulas_file: Path) -> None:
    wb = openpyxl.load_workbook(formulas_file, data_only=False)
    ws = wb.active
    mgmt_net_row = _find_row(ws, LABEL_MGMT_NET)
    for col_idx in _week_cols(ws):
        val = ws.cell(mgmt_net_row, col_idx).value
        assert isinstance(val, str) and val.startswith("="), (
            f"Mgmt Net cell ({mgmt_net_row},{col_idx}) should be a formula, got {val!r}"
        )
        assert "-" in val, f"Mgmt Net formula should subtract absence: {val!r}"


def test_formulas_file_total_weeks_is_sum_formula(formulas_file: Path) -> None:
    wb = openpyxl.load_workbook(formulas_file, data_only=False)
    ws = wb.active
    tw_col = _total_weeks_col(ws)
    for row in ws.iter_rows(min_row=2):
        label = row[1].value
        if label not in _CAPACITY_LABELS and label is not None:
            val = ws.cell(row[0].row, tw_col).value
            assert isinstance(val, str) and val.upper().startswith("=SUM("), (
                f"Total Weeks for '{label}' should be =SUM(...), got {val!r}"
            )


def test_formulas_file_weekly_allocation_is_sum_formula(formulas_file: Path) -> None:
    wb = openpyxl.load_workbook(formulas_file, data_only=False)
    ws = wb.active
    total_row = _find_row(ws, LABEL_TOTAL_ROW)
    for col_idx in _week_cols(ws):
        val = ws.cell(total_row, col_idx).value
        assert isinstance(val, str) and val.upper().startswith("=SUM("), (
            f"Weekly Allocation cell ({total_row},{col_idx}) should be =SUM(...), got {val!r}"
        )


def test_formulas_file_total_row_estimation_is_sum_formula(formulas_file: Path) -> None:
    wb = openpyxl.load_workbook(formulas_file, data_only=False)
    ws = wb.active
    total_row = _find_row(ws, LABEL_TOTAL_ROW)
    # Find Estimation column index
    est_col = next(
        c for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value == OUT_COL_ESTIMATION
    )
    val = ws.cell(total_row, est_col).value
    assert isinstance(val, str) and val.upper().startswith("=SUM("), (
        f"Total row Estimation should be =SUM(...), got {val!r}"
    )


def test_formulas_file_total_row_total_weeks_is_sum_formula(formulas_file: Path) -> None:
    wb = openpyxl.load_workbook(formulas_file, data_only=False)
    ws = wb.active
    total_row = _find_row(ws, LABEL_TOTAL_ROW)
    tw_col = _total_weeks_col(ws)
    val = ws.cell(total_row, tw_col).value
    assert isinstance(val, str) and val.upper().startswith("=SUM("), (
        f"Total row Total Weeks should be =SUM(...), got {val!r}"
    )


def test_values_file_total_row_estimation_is_numeric(values_file: Path) -> None:
    wb = openpyxl.load_workbook(values_file, data_only=True)
    ws = wb.active
    total_row = _find_row(ws, LABEL_TOTAL_ROW)
    est_col = next(
        c for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value == OUT_COL_ESTIMATION
    )
    val = ws.cell(total_row, est_col).value
    assert isinstance(val, (int, float)), (
        f"Total row Estimation should be numeric in values file, got {val!r}"
    )


def test_values_file_total_row_total_weeks_is_numeric(values_file: Path) -> None:
    wb = openpyxl.load_workbook(values_file, data_only=True)
    ws = wb.active
    total_row = _find_row(ws, LABEL_TOTAL_ROW)
    tw_col = _total_weeks_col(ws)
    val = ws.cell(total_row, tw_col).value
    assert isinstance(val, (int, float)), (
        f"Total row Total Weeks should be numeric in values file, got {val!r}"
    )


def test_formulas_eng_net_references_correct_rows(formulas_file: Path) -> None:
    """Spot-check: Eng Net formula references Eng Bruto and Eng Absence rows."""
    wb = openpyxl.load_workbook(formulas_file, data_only=False)
    ws = wb.active
    r_bruto = _find_row(ws, LABEL_ENG_BRUTO)
    r_absence = _find_row(ws, LABEL_ENG_ABSENCE)
    r_net = _find_row(ws, LABEL_ENG_NET)
    week_cols = _week_cols(ws)
    # Check first week column
    col_idx = week_cols[0]
    formula = ws.cell(r_net, col_idx).value
    assert str(r_bruto) in formula, f"Formula {formula!r} should reference bruto row {r_bruto}"
    assert str(r_absence) in formula, f"Formula {formula!r} should reference absence row {r_absence}"


def test_formulas_total_weeks_spans_all_week_columns(formulas_file: Path) -> None:
    """Total Weeks SUM range must start at the first week col and end at the last."""
    wb = openpyxl.load_workbook(formulas_file, data_only=False)
    ws = wb.active
    tw_col = _total_weeks_col(ws)
    week_col_indices = _week_cols(ws)
    first_letter = get_column_letter(week_col_indices[0])
    last_letter = get_column_letter(week_col_indices[-1])
    # Check first epic row
    for row in ws.iter_rows(min_row=2):
        label = row[1].value
        if label not in _CAPACITY_LABELS and label is not None:
            r = row[0].row
            formula = ws.cell(r, tw_col).value
            assert first_letter in formula, f"SUM should start at {first_letter}: {formula!r}"
            assert last_letter in formula, f"SUM should end at {last_letter}: {formula!r}"
            break


# ---------------------------------------------------------------------------
# Parametrized tests: formula positions are correct for 1, 3, and 5 epics
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("n_epics", [1, 3, 5])
def test_weekly_allocation_sum_spans_exact_epic_rows(n_epics: int, tmp_path: Path) -> None:
    """
    The Weekly Allocation SUM formula must reference exactly the epic rows —
    from the row after the last capacity header to the row before the total row.
    This is verified for 1, 3, and 5 epics so that changes in sheet height are caught.
    """
    df = build_output_table(_make_epics(n_epics), _CAPACITY, _START, _END)
    p = tmp_path / f"out_{n_epics}.xlsx"
    n_base = len([c for c in df.columns if c not in {"Budget Bucket", "Epic Description", "Priority", "Estimation", "Total Weeks", "Off Estimate"}])
    write_output_with_formulas(df, p, n_base)

    wb = openpyxl.load_workbook(p, data_only=False)
    ws = wb.active

    # Identify expected epic rows from the sheet directly
    epic_rows_in_sheet = [
        row[0].row
        for row in ws.iter_rows(min_row=2)
        if row[1].value not in _CAPACITY_LABELS and row[1].value is not None
    ]
    assert len(epic_rows_in_sheet) == n_epics, (
        f"Expected {n_epics} epic rows, found {len(epic_rows_in_sheet)}"
    )

    expected_first = epic_rows_in_sheet[0]
    expected_last  = epic_rows_in_sheet[-1]
    total_row_num  = _find_row(ws, LABEL_TOTAL_ROW)

    for col_idx in _week_cols(ws):
        cl = get_column_letter(col_idx)
        formula = ws.cell(total_row_num, col_idx).value
        expected = f"=SUM({cl}{expected_first}:{cl}{expected_last})"
        assert formula == expected, (
            f"n_epics={n_epics}: Weekly Allocation formula for col {cl} "
            f"expected {expected!r}, got {formula!r}"
        )


@pytest.mark.parametrize("n_epics", [1, 3, 5])
def test_total_weeks_sum_references_own_row(n_epics: int, tmp_path: Path) -> None:
    """
    The Total Weeks formula for each epic must reference that epic's own row number
    (not a hardcoded constant), across sheet shapes with 1, 3, and 5 epics.
    """
    df = build_output_table(_make_epics(n_epics), _CAPACITY, _START, _END)
    p = tmp_path / f"out_{n_epics}.xlsx"
    n_base = len([c for c in df.columns if c not in {"Budget Bucket", "Epic Description", "Priority", "Estimation", "Total Weeks", "Off Estimate"}])
    write_output_with_formulas(df, p, n_base)

    wb = openpyxl.load_workbook(p, data_only=False)
    ws = wb.active

    tw_col = _total_weeks_col(ws)
    week_col_indices = _week_cols(ws)
    first_letter = get_column_letter(week_col_indices[0])
    last_letter  = get_column_letter(week_col_indices[-1])

    for row in ws.iter_rows(min_row=2):
        label = row[1].value
        if label in _CAPACITY_LABELS or label is None:
            continue
        r = row[0].row
        formula = ws.cell(r, tw_col).value
        expected = f"=SUM({first_letter}{r}:{last_letter}{r})"
        assert formula == expected, (
            f"n_epics={n_epics}, epic row {r}: "
            f"Total Weeks expected {expected!r}, got {formula!r}"
        )


@pytest.mark.parametrize("n_epics", [1, 3, 5])
def test_net_capacity_formulas_reference_fixed_header_rows(n_epics: int, tmp_path: Path) -> None:
    """
    Eng Net and Mgmt Net formulas must reference the bruto/absence rows, which
    are always at fixed positions (rows 2-7) regardless of how many epics follow.
    """
    df = build_output_table(_make_epics(n_epics), _CAPACITY, _START, _END)
    p = tmp_path / f"out_{n_epics}.xlsx"
    n_base = len([c for c in df.columns if c not in {"Budget Bucket", "Epic Description", "Priority", "Estimation", "Total Weeks", "Off Estimate"}])
    write_output_with_formulas(df, p, n_base)

    wb = openpyxl.load_workbook(p, data_only=False)
    ws = wb.active

    r_bruto        = _find_row(ws, LABEL_ENG_BRUTO)
    r_eng_absence  = _find_row(ws, LABEL_ENG_ABSENCE)
    r_eng_net      = _find_row(ws, LABEL_ENG_NET)
    r_mgmt_cap     = _find_row(ws, LABEL_MGMT_CAPACITY)
    r_mgmt_absence = _find_row(ws, LABEL_MGMT_ABSENCE)
    r_mgmt_net     = _find_row(ws, LABEL_MGMT_NET)

    for col_idx in _week_cols(ws):
        cl = get_column_letter(col_idx)
        eng_formula  = ws.cell(r_eng_net,  col_idx).value
        mgmt_formula = ws.cell(r_mgmt_net, col_idx).value
        assert eng_formula == f"={cl}{r_bruto}-{cl}{r_eng_absence}", (
            f"n_epics={n_epics}: Eng Net formula wrong: {eng_formula!r}"
        )
        assert mgmt_formula == f"={cl}{r_mgmt_cap}-{cl}{r_mgmt_absence}", (
            f"n_epics={n_epics}: Mgmt Net formula wrong: {mgmt_formula!r}"
        )


# ---------------------------------------------------------------------------
# Off Estimate column and Off Capacity row tests
# ---------------------------------------------------------------------------

def test_values_file_off_estimate_column_is_boolean(values_file: Path) -> None:
    wb = openpyxl.load_workbook(values_file, data_only=True)
    ws = wb.active
    off_est_col = next(
        c for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value == OUT_COL_OFF_ESTIMATE
    )
    for row in ws.iter_rows(min_row=2):
        label = row[1].value
        if label not in _CAPACITY_LABELS and label is not None:
            val = ws.cell(row[0].row, off_est_col).value
            assert isinstance(val, bool), (
                f"Off Estimate for '{label}' should be bool, got {val!r}"
            )


def test_values_file_off_capacity_row_is_boolean(values_file: Path) -> None:
    wb = openpyxl.load_workbook(values_file, data_only=True)
    ws = wb.active
    alert_row = _find_row(ws, LABEL_CAPACITY_ALERT_ROW)
    for col_idx in _week_cols(ws):
        val = ws.cell(alert_row, col_idx).value
        assert isinstance(val, bool), (
            f"Off Capacity cell ({alert_row},{col_idx}) should be bool, got {val!r}"
        )


def test_formulas_file_off_estimate_has_abs_formula(formulas_file: Path) -> None:
    wb = openpyxl.load_workbook(formulas_file, data_only=False)
    ws = wb.active
    off_est_col = next(
        c for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value == OUT_COL_OFF_ESTIMATE
    )
    for row in ws.iter_rows(min_row=2):
        label = row[1].value
        if label not in _CAPACITY_LABELS and label is not None:
            val = ws.cell(row[0].row, off_est_col).value
            assert isinstance(val, str) and "ABS" in val.upper(), (
                f"Off Estimate for '{label}' should be =ABS(...) formula, got {val!r}"
            )


def test_formulas_file_off_capacity_has_abs_formula(formulas_file: Path) -> None:
    wb = openpyxl.load_workbook(formulas_file, data_only=False)
    ws = wb.active
    alert_row = _find_row(ws, LABEL_CAPACITY_ALERT_ROW)
    for col_idx in _week_cols(ws):
        val = ws.cell(alert_row, col_idx).value
        assert isinstance(val, str) and "ABS" in val.upper(), (
            f"Off Capacity cell ({alert_row},{col_idx}) should be =ABS(...) formula, got {val!r}"
        )


def _cf_rules_for_range(ws, cell_range: str):
    marker = f"ConditionalFormatting {cell_range}>"
    for sqref, rules in ws.conditional_formatting._cf_rules.items():
        if marker in str(sqref):
            return rules
    return []


def test_values_file_has_conditional_formatting_rules(values_file: Path) -> None:
    wb = openpyxl.load_workbook(values_file, data_only=False)
    ws = wb.active

    week_cols = _week_cols(ws)
    first_week = get_column_letter(week_cols[0])
    last_week = get_column_letter(week_cols[-1])

    off_est_col = next(
        c for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value == OUT_COL_OFF_ESTIMATE
    )
    off_est_letter = get_column_letter(off_est_col)
    bucket_col = next(
        c for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value == OUT_COL_BUDGET_BUCKET
    )
    bucket_col_letter = get_column_letter(bucket_col)

    r_alert = _find_row(ws, LABEL_CAPACITY_ALERT_ROW)
    epic_rows = [
        row[0].row
        for row in ws.iter_rows(min_row=2)
        if row[1].value not in _CAPACITY_LABELS and row[1].value is not None
    ]
    first_epic = epic_rows[0]
    last_epic = epic_rows[-1]

    off_est_range = f"{off_est_letter}{first_epic}:{off_est_letter}{last_epic}"
    off_est_rules = _cf_rules_for_range(ws, off_est_range)
    assert off_est_rules, f"Expected conditional formatting on range {off_est_range}"
    expected_off_est_formula = f"{off_est_letter}{first_epic}=TRUE"
    assert any(
        getattr(rule, "type", "") == "expression"
        and expected_off_est_formula in getattr(rule, "formula", [])
        for rule in off_est_rules
    ), "Missing expression rule for TRUE Off Estimate cells"

    alert_range = f"{first_week}{r_alert}:{last_week}{r_alert}"
    alert_rules = _cf_rules_for_range(ws, alert_range)
    assert alert_rules, f"Expected conditional formatting on range {alert_range}"
    expected_alert_formula = f"{first_week}{r_alert}=TRUE"
    assert any(
        getattr(rule, "type", "") == "expression"
        and expected_alert_formula in getattr(rule, "formula", [])
        for rule in alert_rules
    ), "Missing expression rule for TRUE Off Capacity cells"

    full_data_range = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    full_row_rules = _cf_rules_for_range(ws, full_data_range)
    assert full_row_rules, f"Expected conditional formatting on range {full_data_range}"
    expected_bucket_formulas = {
        f'${bucket_col_letter}2="Self-Service ML EV Range - Phase 1"',
        f'${bucket_col_letter}2="Quality improvements through ML/AI experimentation"',
        f'${bucket_col_letter}2="Maintenance & Release"',
        f'${bucket_col_letter}2="Security & Compliance"',
        f'${bucket_col_letter}2="Customer Support"',
        f'${bucket_col_letter}2="Critical Technical Debt"',
        f'${bucket_col_letter}2="Critical Product Debt"',
        f'${bucket_col_letter}2="Critical Customer Commitments"',
    }
    found_formulas = {
        rule.formula[0]
        for rule in full_row_rules
        if getattr(rule, "type", "") == "expression" and getattr(rule, "formula", None)
    }
    assert expected_bucket_formulas.issubset(found_formulas), (
        "Missing one or more budget-bucket row color rules"
    )


def test_formulas_file_has_conditional_formatting_rules(formulas_file: Path) -> None:
    wb = openpyxl.load_workbook(formulas_file, data_only=False)
    ws = wb.active

    week_cols = _week_cols(ws)
    first_week = get_column_letter(week_cols[0])
    last_week = get_column_letter(week_cols[-1])

    off_est_col = next(
        c for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value == OUT_COL_OFF_ESTIMATE
    )
    off_est_letter = get_column_letter(off_est_col)
    r_alert = _find_row(ws, LABEL_CAPACITY_ALERT_ROW)
    epic_rows = [
        row[0].row
        for row in ws.iter_rows(min_row=2)
        if row[1].value not in _CAPACITY_LABELS and row[1].value is not None
    ]
    first_epic = epic_rows[0]
    last_epic = epic_rows[-1]

    off_est_range = f"{off_est_letter}{first_epic}:{off_est_letter}{last_epic}"
    off_est_rules = _cf_rules_for_range(ws, off_est_range)
    assert off_est_rules, (
        f"Expected conditional formatting on range {off_est_range}"
    )
    assert any(
        getattr(rule, "type", "") == "expression"
        and f"{off_est_letter}{first_epic}=TRUE" in getattr(rule, "formula", [])
        for rule in off_est_rules
    )

    alert_range = f"{first_week}{r_alert}:{last_week}{r_alert}"
    alert_rules = _cf_rules_for_range(ws, alert_range)
    assert alert_rules, (
        f"Expected conditional formatting on range {alert_range}"
    )
    assert any(
        getattr(rule, "type", "") == "expression"
        and f"{first_week}{r_alert}=TRUE" in getattr(rule, "formula", [])
        for rule in alert_rules
    )


# ---------------------------------------------------------------------------
# Per-week capacity: read_input with D.M. week columns
# ---------------------------------------------------------------------------

_PER_WEEK_FIXTURE = Path(__file__).parent / "data" / "input_per_week_absence.xlsx"

_Q2_MONDAYS = [
    date(2026, 3, 30) + timedelta(weeks=i) for i in range(13)
]


def test_per_week_fixture_loads_bruto_by_week(tmp_path: Path) -> None:
    """Per-week bruto is extracted from D.M. columns when all 13 Q2 weeks present."""
    _, cap = read_input(_PER_WEEK_FIXTURE, 2)
    assert cap.eng_bruto_by_week is not None
    assert len(cap.eng_bruto_by_week) == 13
    # first 6 weeks → 2.0, last 7 weeks → 3.0
    for monday in _Q2_MONDAYS[:6]:
        assert cap.eng_bruto_by_week[monday] == pytest.approx(2.0)
    for monday in _Q2_MONDAYS[6:]:
        assert cap.eng_bruto_by_week[monday] == pytest.approx(3.0)


def test_per_week_fixture_loads_absence_by_week(tmp_path: Path) -> None:
    """Per-week absence is extracted; NaN weeks default to 0."""
    _, cap = read_input(_PER_WEEK_FIXTURE, 2)
    assert cap.eng_absence_by_week is not None
    assert len(cap.eng_absence_by_week) == 13
    # NaN weeks (index 4, 9, 11) default to 0
    nan_indices = [4, 9, 11]
    for i in nan_indices:
        assert cap.eng_absence_by_week[_Q2_MONDAYS[i]] == pytest.approx(0.0)


def test_per_week_eng_net_for_varies_by_week(tmp_path: Path) -> None:
    """eng_net_for returns different values for different weeks."""
    _, cap = read_input(_PER_WEEK_FIXTURE, 2)
    # week 0: bruto=2.0, absence=0.1 → net=1.9
    assert cap.eng_net_for(_Q2_MONDAYS[0]) == pytest.approx(1.9)
    # week 6: bruto=3.0, absence=0.0 → net=3.0
    assert cap.eng_net_for(_Q2_MONDAYS[6]) == pytest.approx(3.0)


def test_per_week_build_output_uses_variable_capacity(tmp_path: Path) -> None:
    """build_output_table uses per-week capacity rows in its output."""
    from planzen.core_logic import build_output_table
    from planzen.config import LABEL_ENG_BRUTO, FISCAL_QUARTERS
    epics_df, cap = read_input(_PER_WEEK_FIXTURE, 2)
    start, end = FISCAL_QUARTERS[2]
    output = build_output_table(epics_df, cap, start, end)
    bruto_row = output[output["Epic Description"] == LABEL_ENG_BRUTO].iloc[0]
    week_cols = [c for c in output.columns if c not in
                 {"Budget Bucket", "Epic Description", "Priority", "Estimation", "Total Weeks", "Off Estimate"}]
    # first 6 weeks should be 2.0, last 7 should be 3.0
    for i, w in enumerate(week_cols[:6]):
        assert bruto_row[w] == pytest.approx(2.0), f"Week {w} bruto should be 2.0"
    for i, w in enumerate(week_cols[6:]):
        assert bruto_row[w] == pytest.approx(3.0), f"Week {w} bruto should be 3.0"


def test_partial_per_week_bruto_raises_error(tmp_path: Path) -> None:
    """If only some Q2 weeks have bruto values, a ValueError is raised."""
    from datetime import date, timedelta
    q2_mondays = [date(2026, 3, 30) + timedelta(weeks=i) for i in range(13)]
    week_cols = [f"{m.day}.{m.month}." for m in q2_mondays]
    rows = [
        {"Budget Bucket": "Num Engineers", "Estimation": 3.0,
         "Epic Description": "", "Priority": "", "Link": "", "Type": ""},
    ]
    # Engineer Capacity (Bruto) with only 5 of 13 Q2 weeks
    bruto_row = {"Budget Bucket": "Engineer Capacity (Bruto)", "Estimation": "",
                 "Epic Description": "", "Priority": "", "Link": "", "Type": ""}
    for w in week_cols[:5]:
        bruto_row[w] = 3.0
    rows.append(bruto_row)
    rows.append({"Budget Bucket": "Platform", "Epic Description": "X", "Estimation": 2.0,
                 "Priority": 1, "Link": "http://x", "Type": "Feature"})
    p = tmp_path / "partial.xlsx"
    pd.DataFrame(rows).to_excel(p, index=False)
    with pytest.raises(ValueError, match="partially populated"):
        read_input(p, 2)


# ---------------------------------------------------------------------------
# Design choices: config detection, Priority imputation, unnamed columns
# ---------------------------------------------------------------------------

def test_config_row_detected_by_epic_description_column(tmp_path: Path) -> None:
    """A row with 'Engineer Capacity (Bruto)' in Epic Description is a config row."""
    p = tmp_path / "input.xlsx"
    rows = [
        # Config rows identified via Epic Description (no Budget Bucket on these rows)
        {"Epic Description": "Engineer Capacity (Bruto)", "Estimation": 4.0},
        {"Epic Description": "Management Capacity (Bruto)", "Estimation": 1.0},
        _base_row(),
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    _, cap = read_input(p, 2)
    assert cap.num_engineers == 4.0
    assert cap.num_managers == 1.0


def test_config_row_epic_description_case_insensitive(tmp_path: Path) -> None:
    """Config label in Epic Description is matched case-insensitively."""
    p = tmp_path / "input.xlsx"
    rows = [
        {"Epic Description": "engineer capacity (bruto)", "Estimation": 3.0},
        _base_row(),
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    _, cap = read_input(p, 2)
    assert cap.num_engineers == 3.0


def test_config_row_not_included_in_epics(tmp_path: Path) -> None:
    """Config rows identified via Epic Description must not appear as epics."""
    p = tmp_path / "input.xlsx"
    rows = [
        {"Epic Description": "Engineer Capacity (Bruto)", "Estimation": 4.0},
        _base_row(),
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    epics, _ = read_input(p, 2)
    epic_descs = list(epics["Epic Description"])
    assert "Engineer Capacity (Bruto)" not in epic_descs
    assert len(epics) == 1


def test_validate_passes_when_all_config_in_epic_description(tmp_path: Path) -> None:
    """validate_input_file returns no errors when config labels are in Epic Description."""
    p = tmp_path / "input.xlsx"
    rows = [
        {"Epic Description": "Engineer Capacity (Bruto)", "Estimation": 5.0, "Budget Bucket": None},
        {"Epic Description": "Management Capacity (Bruto)", "Estimation": 2.0, "Budget Bucket": None},
        {"Epic Description": "Engineer Absence", "Estimation": 10.0, "Budget Bucket": None},
        _base_row(**{"Epic Description": "Epic X"}),
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    errors = validate_input_file(p)
    assert errors == [], f"Unexpected errors: {errors}"


def test_read_input_absence_in_epic_description(tmp_path: Path) -> None:
    """Engineer Absence label in Epic Description is read as absence config."""
    p = tmp_path / "input.xlsx"
    rows = [
        {"Epic Description": "Engineer Capacity (Bruto)", "Estimation": 5.0, "Budget Bucket": None},
        {"Epic Description": "Engineer Absence", "Estimation": 15.0, "Budget Bucket": None},
        _base_row(),
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    _, cap = read_input(p, 2)
    # 15 absence days / (5 days/week * 13 weeks) ≈ 0.231 PW/week
    import math
    assert math.isclose(cap.eng_absence_per_week, 15.0 / 5 / 13, rel_tol=1e-6)


def test_read_input_num_engineers_in_epic_description(tmp_path: Path) -> None:
    """Num Engineers label in Epic Description is accepted as headcount config."""
    p = tmp_path / "input.xlsx"
    rows = [
        {"Epic Description": "Num Engineers", "Estimation": 6.0, "Budget Bucket": None},
        _base_row(),
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    _, cap = read_input(p, 2)
    assert cap.num_engineers == 6.0


def test_read_input_mgmt_absence_in_epic_description(tmp_path: Path) -> None:
    """Management Absence label in Epic Description is read as mgmt absence config."""
    p = tmp_path / "input.xlsx"
    rows = [
        {"Epic Description": "Engineer Capacity (Bruto)", "Estimation": 5.0, "Budget Bucket": None},
        {"Epic Description": "Management Capacity (Bruto)", "Estimation": 2.0, "Budget Bucket": None},
        {"Epic Description": "Management Absence", "Estimation": 5.0, "Budget Bucket": None},
        _base_row(),
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    _, cap = read_input(p, 2)
    import math
    assert math.isclose(cap.mgmt_absence_per_week, 5.0 / 5 / 13, rel_tol=1e-6)


def test_config_in_epic_description_coexists_with_budget_bucket_config(tmp_path: Path) -> None:
    """Config rows may mix Epic Description and Budget Bucket columns for labels."""
    p = tmp_path / "input.xlsx"
    rows = [
        # Engineer Bruto via Epic Description
        {"Epic Description": "Engineer Capacity (Bruto)", "Estimation": 4.0, "Budget Bucket": None},
        # Management via Budget Bucket (classic format)
        {"Budget Bucket": "Management Capacity (Bruto)", "Estimation": 3.0},
        _base_row(),
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    _, cap = read_input(p, 2)
    assert cap.num_engineers == 4.0
    assert cap.num_managers == 3.0


def test_rows_without_budget_bucket_are_dropped(tmp_path: Path) -> None:
    """Rows with an Epic Description but no Budget Bucket are silently dropped."""
    p = tmp_path / "input.xlsx"
    rows = _config_rows() + [
        _base_row(**{"Epic Description": "Real Epic"}),
        # Annotation / computed row: has Epic Description, no Budget Bucket
        {"Epic Description": "Holiday team members", "Estimation": None},
        {"Epic Description": "Capacity incl. EM"},
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    epics, _ = read_input(p, 2)
    assert list(epics["Epic Description"]) == ["Real Epic"]


def test_unnamed_columns_are_dropped_silently(tmp_path: Path) -> None:
    """Columns whose header is 'Unnamed: N' (pandas default) are silently dropped."""
    import openpyxl as ox
    p = tmp_path / "input.xlsx"
    _write_input(p, [_base_row()])
    # Inject an unnamed column by writing a value with no header
    wb = ox.load_workbook(p)
    ws = wb.active
    last_col = ws.max_column + 1
    ws.cell(1, last_col).value = None   # no header
    ws.cell(2, last_col).value = "noise"
    wb.save(p)
    epics, _ = read_input(p, 2)
    assert not any(str(c).startswith("Unnamed") for c in epics.columns)


def test_priority_imputed_from_bucket_when_column_absent(tmp_path: Path) -> None:
    """When the Priority column is absent, Priority is imputed from Budget Bucket."""
    from planzen.config import BUCKET_PRIORITY
    p = tmp_path / "input.xlsx"
    bucket = "Customer Support"
    rows = _config_rows() + [
        {"Epic Description": "E", "Estimation": 1.0, "Budget Bucket": bucket, "Link": "x"},
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    epics, _ = read_input(p, 2)
    assert int(epics.iloc[0]["Priority"]) == BUCKET_PRIORITY[bucket]


def test_priority_imputed_from_bucket_when_value_blank(tmp_path: Path) -> None:
    """A blank Priority cell is filled from Budget Bucket."""
    from planzen.config import BUCKET_PRIORITY
    p = tmp_path / "input.xlsx"
    bucket = "Security & Compliance"
    row = _base_row(**{"Budget Bucket": bucket})
    row["Priority"] = None
    _write_input(p, [row])
    epics, _ = read_input(p, 2)
    assert int(epics.iloc[0]["Priority"]) == BUCKET_PRIORITY[bucket]


def test_priority_defaults_to_999_for_unknown_bucket(tmp_path: Path) -> None:
    """Unknown Budget Bucket → Priority defaults to 999 (lowest)."""
    p = tmp_path / "input.xlsx"
    row = _base_row(**{"Budget Bucket": "Some Unknown Bucket"})
    row["Priority"] = None
    _write_input(p, [row])
    epics, _ = read_input(p, 2)
    assert int(epics.iloc[0]["Priority"]) == 999


def test_explicit_priority_is_preserved(tmp_path: Path) -> None:
    """Explicitly set Priority values are never overwritten by imputation."""
    p = tmp_path / "input.xlsx"
    row = _base_row(**{"Budget Bucket": "Customer Support", "Priority": 5})
    _write_input(p, [row])
    epics, _ = read_input(p, 2)
    assert int(epics.iloc[0]["Priority"]) == 5


def test_validate_no_error_for_missing_priority_column(tmp_path: Path) -> None:
    """Missing Priority column is not a validation error (imputed from Budget Bucket)."""
    p = tmp_path / "input.xlsx"
    rows = _config_rows() + [
        {"Epic Description": "E", "Estimation": 1.0, "Budget Bucket": "Customer Support"},
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    errors = validate_input_file(p)
    assert not any("Priority" in e for e in errors)


def test_extra_unrecognized_columns_do_not_cause_errors(tmp_path: Path) -> None:
    """Extra columns (helper totals, notes, etc.) are silently ignored."""
    p = tmp_path / "input.xlsx"
    rows = _config_rows() + [
        _base_row(**{
            "Some helper total": 42,
            "Q1 notes": "carry over",
            "Old estimation": 99,
        }),
    ]
    pd.DataFrame(rows).to_excel(p, index=False)
    errors = validate_input_file(p)
    assert errors == []
    epics, _ = read_input(p, 2)
    assert len(epics) == 1


def test_per_week_only_bruto_no_scalar(tmp_path: Path) -> None:
    """Per-week-only mode: Engineer Capacity (Bruto) has week values, no scalar Estimation.

    Models the real-world case where the config row has values only in the week
    columns and no value in the Estimation column.
    """
    q2_mondays = [date(2026, 3, 30) + timedelta(weeks=i) for i in range(13)]
    week_cols = [f"{m.day}.{m.month}." for m in q2_mondays]

    bruto_row: dict = {
        "Budget Bucket": "Engineer Capacity (Bruto)",
        "Epic Description": "Engineer Capacity (Bruto)",
        "Estimation": None,
    }
    for w in week_cols:
        bruto_row[w] = 2.0

    epic_row: dict = {
        "Budget Bucket": "Customer Support",
        "Epic Description": "Fix production bug",
        "Estimation": 3.0,
        "Priority": 0,
    }

    p = tmp_path / "per_week_only.xlsx"
    pd.DataFrame([bruto_row, epic_row]).to_excel(p, index=False)

    # validate should not flag missing engineer capacity
    errors = validate_input_file(p, quarter=2)
    assert not any("Missing engineer capacity" in e for e in errors), errors

    # read_input should succeed and populate eng_bruto_by_week
    _, cap = read_input(p, 2)
    assert cap.eng_bruto_by_week is not None
    assert len(cap.eng_bruto_by_week) == 13
    for v in cap.eng_bruto_by_week.values():
        assert v == pytest.approx(2.0)
    # scalar fallback derived from mean of per-week values
    assert cap.num_engineers == pytest.approx(2.0)
