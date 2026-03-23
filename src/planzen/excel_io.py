"""
Excel I/O for planzen.

Reads the input plan file and writes the output allocation table.
All file operations are isolated here; core_logic stays pure.
"""

from __future__ import annotations

import logging
from datetime import date, datetime as _datetime, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from planzen.config import (
    ALLOC_MODE_DEFAULT,
    BUCKET_COLORS,
    BUCKET_PRIORITY,
    COL_ALLOC_MODE,
    COL_BUDGET_BUCKET,
    COL_EPIC,
    COL_ESTIMATION,
    COL_LINK,
    COL_PRIORITY,
    DEFAULT_MGMT_CAPACITY_PW,
    FISCAL_QUARTERS,
    TEAM_CONFIG_LABELS,
    TEAM_LABEL_ENG_ABSENCE,
    TEAM_LABEL_ENGINEERS,
    TEAM_LABEL_MGMT_ABSENCE,
    TEAM_LABEL_MANAGERS,
    TEAM_LABEL_NUM_ENGINEERS,
    LABEL_CAPACITY_ALERT_ROW,
    LABEL_ENG_ABSENCE,
    LABEL_ENG_BRUTO,
    LABEL_ENG_NET,
    LABEL_MGMT_ABSENCE,
    LABEL_MGMT_CAPACITY,
    LABEL_MGMT_NET,
    LABEL_TOTAL_ROW,
    OUT_COL_BUDGET_BUCKET,
    OUT_COL_EPIC,
    OUT_COL_ESTIMATION,
    OUT_COL_OFF_ESTIMATE,
    OUT_COL_PRIORITY,
    OUT_COL_TOTAL_WEEKS,
    VALID_ALLOC_MODES,
)
from planzen.core_logic import CapacityConfig

# Columns that must be present in the input file.
# Priority is intentionally absent — it is imputed from Budget Bucket when missing.
# Link is optional.
REQUIRED_INPUT_COLUMNS = {
    COL_EPIC,
    COL_ESTIMATION,
    COL_BUDGET_BUCKET,
}

_log = logging.getLogger(__name__)

_NON_WEEK_COLS = {
    OUT_COL_BUDGET_BUCKET, OUT_COL_EPIC, OUT_COL_PRIORITY,
    OUT_COL_ESTIMATION, OUT_COL_TOTAL_WEEKS, OUT_COL_OFF_ESTIMATE,
}

_CAPACITY_LABELS = {
    LABEL_ENG_BRUTO, LABEL_ENG_ABSENCE, LABEL_ENG_NET,
    LABEL_MGMT_CAPACITY, LABEL_MGMT_ABSENCE, LABEL_MGMT_NET,
    LABEL_TOTAL_ROW, LABEL_CAPACITY_ALERT_ROW,
}

_ALERT_FILL = PatternFill(fill_type="solid", start_color="00FFC7CE", end_color="00FFC7CE")
_ALERT_FONT = Font(color="009C0006")
# Sourced from config so both CLI and web share the same mapping.
_BUCKET_COLOR_BY_LABEL = BUCKET_COLORS


def _normalize_config_label(s: object) -> str:
    """Normalise a config row label for fuzzy matching.

    Applies in order:
    1. Strip whitespace and lowercase.
    2. Remove parenthetical suffixes, e.g. ``(days)``.
    3. Singularize common plural words (trailing *s* on words longer than 3
       characters) so ``"Engineers"`` and ``"Engineer"`` both match.
    """
    if not isinstance(s, str):
        return ""
    import re
    s = s.strip().lower()
    s = re.sub(r"\s*\(.*?\)", "", s).strip()
    s = " ".join(
        w[:-1] if w.endswith("s") and len(w) > 3 else w
        for w in s.split()
    )
    return s


# Mapping: normalised canonical label → canonical label string
_TEAM_CONFIG_LABELS_NORM: dict[str, str] = {
    _normalize_config_label(label): label
    for label in TEAM_CONFIG_LABELS
}

_KNOWN_COLUMNS: set[str] = {
    COL_EPIC, COL_ESTIMATION, COL_BUDGET_BUCKET, COL_LINK,
    COL_PRIORITY, COL_ALLOC_MODE,
}
_KNOWN_COLUMNS_LOWER: dict[str, str] = {c.lower(): c for c in _KNOWN_COLUMNS}


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename DataFrame columns to their canonical form (case-insensitive, stripped)."""
    rename = {
        col: _KNOWN_COLUMNS_LOWER[col.strip().lower()]
        for col in df.columns
        if isinstance(col, str) and col.strip().lower() in _KNOWN_COLUMNS_LOWER
        and col != _KNOWN_COLUMNS_LOWER[col.strip().lower()]
    }
    if rename:
        _log.info("Normalising column names: %s", rename)
    return df.rename(columns=rename)


def _drop_unnamed_columns(df: pd.DataFrame, keep: set[str] | None = None) -> pd.DataFrame:
    """Drop columns with no header (pandas names them 'Unnamed: N').

    Columns whose names are in *keep* are preserved even if unnamed.
    """
    preserve = keep or set()
    unnamed = [
        c for c in df.columns
        if isinstance(c, str) and c.startswith("Unnamed:") and c not in preserve
    ]
    if unnamed:
        _log.debug("Dropping %d unnamed column(s): %s", len(unnamed), unnamed)
    return df.drop(columns=unnamed)


def formulas_path(path: Path) -> Path:
    """Return the sibling path for the formulas variant of an output file."""
    return path.with_stem(path.stem + "_formulas")


def _config_label_series(df: pd.DataFrame) -> pd.Series:
    """Return a Series of normalised config labels for each row.

    Checks ``Budget Bucket`` first; falls back to ``Type``, then to
    ``Epic Description`` when neither of the first two carries a recognised
    config label.  This allows files where config rows are labelled only in
    the Epic Description column (rather than in Budget Bucket / Type).
    """
    COL_TYPE = "Type"

    def _norm_col(col: str) -> pd.Series:
        if col in df.columns:
            return df[col].fillna("").astype(str).str.strip().apply(_normalize_config_label)
        return pd.Series("", index=df.index)

    bucket_norm = _norm_col(COL_BUDGET_BUCKET)
    type_norm   = _norm_col(COL_TYPE)
    epic_norm   = _norm_col(COL_EPIC)

    result = bucket_norm
    result = result.where(result.isin(_TEAM_CONFIG_LABELS_NORM), type_norm)
    result = result.where(result.isin(_TEAM_CONFIG_LABELS_NORM), epic_norm)
    return result


def _quarter_mondays(quarter: int) -> list[date]:
    """Return all 13 Mondays in the given fiscal quarter."""
    start, end = FISCAL_QUARTERS[quarter]
    day, mondays = start, []
    while day <= end:
        mondays.append(day)
        day += timedelta(weeks=1)
    return mondays


def _week_col_candidates(monday: date) -> list[str]:
    """Return all recognised column-name formats for a given Monday.

    Supported formats:
    * ``D.M.``   – e.g. ``30.3.``
    * ``D-Mon``  – e.g. ``30-Mar``
    """
    return [
        f"{monday.day}.{monday.month}.",
        f"{monday.day}-{monday.strftime('%b')}",
    ]


def _parse_dm_week_columns(
    df_columns: pd.Index,
    quarter_mondays: list[date],
) -> dict[date, str]:
    """Map each quarter Monday to its week column name.

    Recognises ``D.M.`` (e.g. ``30.3.``) and ``D-Mon`` (e.g. ``30-Mar``)
    formats.  Only returns entries where the column actually exists in
    *df_columns*.
    """
    col_set = {str(c) for c in df_columns}
    result: dict[date, str] = {}
    for monday in quarter_mondays:
        for candidate in _week_col_candidates(monday):
            if candidate in col_set:
                result[monday] = candidate
                break
    return result


def _find_week_columns_in_data(
    df: pd.DataFrame,
    quarter_mondays: list[date],
) -> tuple[dict[date, str], int] | tuple[None, None]:
    """Scan data rows for one containing datetime values matching the quarter Mondays.

    Some files place week dates as cell values in a dedicated row (e.g. row 3)
    rather than as column headers.  Accepts a row that contains at least
    ``len(quarter_mondays) - 1`` matching dates to tolerate a single missing
    week at the end of the quarter.  Returns ``(date→column_name, row_index)``
    when such a row is found, otherwise ``(None, None)``.
    """
    monday_set = set(quarter_mondays)
    min_match = len(quarter_mondays) - 1  # tolerate one missing week
    for idx, row in df.iterrows():
        col_for_date: dict[date, str] = {}
        for col in df.columns:
            val = row[col]
            if isinstance(val, pd.Timestamp):
                d = val.date()
            elif isinstance(val, _datetime):
                d = val.date()
            else:
                continue
            if d in monday_set:
                col_for_date[d] = str(col)
        if len(col_for_date) >= min_match:
            _log.info(
                "Week columns detected from data row %d (%d/%d Q dates present).",
                idx, len(col_for_date), len(quarter_mondays),
            )
            return col_for_date, int(idx)
    return None, None


def _load_df(
    path: Path,
    quarter: int | None,
) -> tuple[pd.DataFrame, dict[date, str]]:
    """Read the Excel file, detect week columns, and return a clean DataFrame.

    Week columns are identified first from column headers (``D.M.`` or
    ``D-Mon`` format), then by scanning data rows for datetime values that
    match the quarter Mondays.  A data row used as week headers is dropped
    from the returned DataFrame.  Unnamed columns that are not week columns
    are dropped.

    Returns ``(df, week_col_map)`` where *week_col_map* maps each quarter
    Monday to its column name (empty dict when *quarter* is ``None`` or no
    week columns are found).
    """
    df = pd.read_excel(path)
    df = _normalize_columns(df)

    week_col_map: dict[date, str] = {}
    date_row_idx: int | None = None

    if quarter is not None:
        q_mondays = _quarter_mondays(quarter)
        week_col_map = _parse_dm_week_columns(df.columns, q_mondays)
        if not week_col_map:
            found_map, date_row_idx = _find_week_columns_in_data(df, q_mondays)
            if found_map:
                week_col_map = found_map

    # Drop unnamed columns that are not part of the week column set
    df = _drop_unnamed_columns(df, keep=set(week_col_map.values()))

    # Remove the date-header data row (it is not an epic or config row)
    if date_row_idx is not None:
        df = df.drop(index=date_row_idx).reset_index(drop=True)

    return df, week_col_map


def validate_input_file(path: Path, quarter: int | None = None) -> list[str]:
    """
    Validate the input Excel file and return a list of human-readable error
    messages.  An empty list means the file is valid.

    All issues are collected before returning so the user sees everything at
    once rather than fixing one problem at a time.
    """
    errors: list[str] = []

    # --- file-level ---
    if not path.exists():
        errors.append(
            f"File not found: '{path}'\n"
            "  → Check the path and make sure the file exists."
        )
        return errors  # nothing else can be checked

    try:
        df, week_col_map = _load_df(path, quarter)
    except Exception as exc:
        errors.append(
            f"Cannot read '{path}' as an Excel file: {exc}\n"
            "  → Make sure the file is a valid .xlsx workbook and is not open in Excel."
        )
        return errors

    # --- required structural columns ---
    for col in (COL_EPIC, COL_ESTIMATION):
        if col not in df.columns:
            errors.append(
                f"Missing required column: \"{col}\"\n"
                f"  → Add a column with this exact header name."
            )

    if errors:
        return errors  # can't proceed without these two columns

    # --- team config rows (normalised matching) ---
    epic_norm = _config_label_series(df)
    config_mask = epic_norm.isin(_TEAM_CONFIG_LABELS_NORM)
    config_rows = df[config_mask].copy()
    config_rows[COL_EPIC] = epic_norm[config_mask].map(_TEAM_CONFIG_LABELS_NORM)
    config_df = config_rows.set_index(COL_EPIC)[COL_ESTIMATION]

    # Engineer capacity: week columns (priority 1) → scalar Estimation (priority 2)
    # → Num Engineers (priority 3).
    has_eng_bruto = TEAM_LABEL_ENGINEERS in config_df.index and pd.notna(config_df[TEAM_LABEL_ENGINEERS])
    has_num_eng   = TEAM_LABEL_NUM_ENGINEERS in config_df.index and pd.notna(config_df[TEAM_LABEL_NUM_ENGINEERS])
    # Per-week values in week columns satisfy engineer capacity (partial is accepted with a warning).
    has_per_week_bruto = False
    if week_col_map:
        eng_rows = config_rows[config_rows[COL_EPIC] == TEAM_LABEL_ENGINEERS]
        if not eng_rows.empty:
            eng_row = eng_rows.iloc[0]
            n_with_value = sum(
                1 for col in week_col_map.values()
                if col in eng_row.index and pd.notna(eng_row[col])
            )
            if n_with_value > 0:
                has_per_week_bruto = True  # partial is allowed; missing weeks use scalar fallback

    if not has_per_week_bruto and not has_eng_bruto and not has_num_eng:
        errors.append(
            f'Missing engineer capacity config.\n'
            f'  → Add a row with "{TEAM_LABEL_ENGINEERS}" in the "{COL_EPIC}" or "{COL_BUDGET_BUCKET}" column\n'
            f'    and values in week columns, a numeric Estimation, or use "{TEAM_LABEL_NUM_ENGINEERS}" for headcount.'
        )
    elif not has_per_week_bruto and has_eng_bruto:
        val = config_df[TEAM_LABEL_ENGINEERS]
        try:
            if float(val) <= 0:
                errors.append(f'"{TEAM_LABEL_ENGINEERS}" must be greater than 0 (got {val!r}).')
        except (TypeError, ValueError):
            errors.append(f'"{TEAM_LABEL_ENGINEERS}" must be a number (got {val!r}).')
    elif not has_per_week_bruto and not has_eng_bruto and has_num_eng:
        val = config_df[TEAM_LABEL_NUM_ENGINEERS]
        try:
            if float(val) <= 0:
                errors.append(f'"{TEAM_LABEL_NUM_ENGINEERS}" must be greater than 0 (got {val!r}).')
        except (TypeError, ValueError):
            errors.append(f'"{TEAM_LABEL_NUM_ENGINEERS}" must be a number (got {val!r}).')

    # Management capacity: optional — defaults to DEFAULT_MGMT_CAPACITY_PW if absent
    if TEAM_LABEL_MANAGERS in config_df.index and pd.notna(config_df[TEAM_LABEL_MANAGERS]):
        val = config_df[TEAM_LABEL_MANAGERS]
        try:
            if float(val) <= 0:
                errors.append(f'"{TEAM_LABEL_MANAGERS}" must be greater than 0 (got {val!r}).')
        except (TypeError, ValueError):
            errors.append(f'"{TEAM_LABEL_MANAGERS}" must be a number (got {val!r}).')

    for label in (TEAM_LABEL_ENG_ABSENCE, TEAM_LABEL_MGMT_ABSENCE):
        if label in config_df.index and pd.notna(config_df[label]):
            val = config_df[label]
            try:
                fval = float(val)
                if fval < 0:
                    errors.append(
                        f'"{label}" cannot be negative (got {val!r}).\n'
                        "  → Enter the total absence days for the quarter, e.g. 10."
                        " Use 0 if none, or omit the row to use the default."
                    )
            except (TypeError, ValueError):
                errors.append(
                    f'"{label}" must be a number (got {val!r}).\n'
                    "  → Enter the total absence days for the quarter, e.g. 10."
                )

    # --- epic rows --- (apply same filtering as read_input)
    epics_df = df[~config_mask].reset_index(drop=True)
    epics_df = epics_df.dropna(how="all").reset_index(drop=True)
    epics_df = epics_df[epics_df[COL_EPIC].notna()].reset_index(drop=True)
    # Rows without a Budget Bucket are annotations/computed rows, not epics.
    if COL_BUDGET_BUCKET in epics_df.columns:
        epics_df = epics_df[epics_df[COL_BUDGET_BUCKET].notna()].reset_index(drop=True)

    missing_cols = (REQUIRED_INPUT_COLUMNS - {COL_EPIC, COL_ESTIMATION}) - set(epics_df.columns)
    if missing_cols:
        for col in sorted(missing_cols):
            errors.append(
                f'Missing required epic column: "{col}"\n'
                "  → Add a column with this exact header name."
            )

    if len(epics_df) == 0:
        errors.append(
            "No epic rows found after the config rows.\n"
            "  → Add at least one epic row below the team config rows."
        )
        return errors

    # --- per-row epic data ---
    if COL_ESTIMATION in epics_df.columns:
        for i, row in epics_df.iterrows():
            val = row[COL_ESTIMATION]
            name = row.get(COL_EPIC, f"row {i + 1}")
            if pd.isna(val):
                pass  # empty Estimation defaults to 0 at read time
            else:
                try:
                    fval = float(val)
                    if fval < 0:
                        errors.append(
                            f'Epic "{name}": "Estimation" cannot be negative (got {val!r}).\n'
                            "  → Enter the total effort in Person-Weeks, e.g. 4.5."
                        )
                except (TypeError, ValueError):
                    errors.append(
                        f'Epic "{name}": "Estimation" must be a number (got {val!r}).\n'
                        "  → Enter the total effort in Person-Weeks, e.g. 4.5."
                    )

    # Priority: only validate explicit non-null values; blank values are imputed at read time.
    if COL_PRIORITY in epics_df.columns:
        for i, row in epics_df.iterrows():
            val = row[COL_PRIORITY]
            if pd.notna(val):
                try:
                    float(val)
                except (TypeError, ValueError):
                    name = row.get(COL_EPIC, f"row {i + 1}")
                    errors.append(
                        f'Epic "{name}": "Priority" must be a number (got {val!r}).\n'
                        "  → Enter a numeric priority, e.g. 1."
                    )

    if COL_ALLOC_MODE in epics_df.columns:
        for i, row in epics_df.iterrows():
            val = row[COL_ALLOC_MODE]
            if pd.notna(val) and str(val).strip() and str(val).strip() not in VALID_ALLOC_MODES:
                name = row.get(COL_EPIC, f"row {i + 1}")
                errors.append(
                    f'Epic "{name}": invalid "Allocation Mode": {val!r}.\n'
                    f'  → Valid values: {", ".join(sorted(VALID_ALLOC_MODES))}'
                    f" (or leave blank to use the default: {ALLOC_MODE_DEFAULT})."
                )

    return errors


def read_input(path: Path, quarter: int) -> tuple[pd.DataFrame, CapacityConfig]:
    """
    Read the input Excel file and return epics plus a fully-constructed
    ``CapacityConfig`` for the requested quarter.

    Capacity resolution order (engineer bruto):

    1. **Per-week** — week columns (``D.M.``, ``D-Mon``, or datetime values in
       a data row) carrying per-week bruto values.  All Q-weeks must be
       populated (all-or-nothing); absence is lenient (missing weeks → 0).
    2. **Scalar** — ``Engineer Capacity (Bruto)`` row with a numeric
       ``Estimation`` value.
    3. **Headcount** — ``Num Engineers`` row with a numeric ``Estimation``
       value; 1 PW per FTE assumed.

    Raises
    ------
    ValueError
        If required config rows or epic columns are missing, or if
        per-week bruto is only partially populated for the quarter.
    """
    df, week_col_map = _load_df(path, quarter)

    if COL_EPIC not in df.columns or COL_ESTIMATION not in df.columns:
        raise ValueError(
            f"Input file must have '{COL_EPIC}' and '{COL_ESTIMATION}' columns."
        )

    # --- extract team config rows (normalised matching) ---
    epic_norm = _config_label_series(df)
    config_mask = epic_norm.isin(_TEAM_CONFIG_LABELS_NORM)
    config_rows = df[config_mask].copy()
    config_rows[COL_EPIC] = epic_norm[config_mask].map(_TEAM_CONFIG_LABELS_NORM)
    config_df = config_rows.set_index(COL_EPIC)[COL_ESTIMATION]

    q_mondays = _quarter_mondays(quarter)
    n_weeks = len(q_mondays)

    # --- per-week capacity (priority 1) ---
    eng_bruto_by_week: dict[date, float] | None = None
    eng_absence_by_week: dict[date, float] | None = None

    if week_col_map:
        # Engineer Capacity (Bruto) — all-or-nothing per quarter
        eng_bruto_rows = config_rows[config_rows[COL_EPIC] == TEAM_LABEL_ENGINEERS]
        if not eng_bruto_rows.empty:
            eng_row = eng_bruto_rows.iloc[0]
            per_week = {
                monday: float(eng_row[col])
                for monday, col in week_col_map.items()
                if col in eng_row.index and pd.notna(eng_row[col])
            }
            if len(per_week) == len(q_mondays):
                eng_bruto_by_week = per_week
            elif len(per_week) > 0:
                # Partial per-week data: use what's there; scalar fallback fills the rest.
                missing = [
                    f"{m.day}-{m.strftime('%b')}"
                    for m in q_mondays
                    if m not in per_week
                ]
                _log.warning(
                    "'%s' is partially populated for Q%d: %d/%d weeks have values "
                    "(missing: %s). Missing weeks will use the scalar fallback.",
                    TEAM_LABEL_ENGINEERS, quarter, len(per_week), len(q_mondays),
                    ", ".join(missing),
                )
                eng_bruto_by_week = per_week

        # Engineer Absence — lenient: missing/NaN weeks default to 0
        eng_absence_rows = config_rows[config_rows[COL_EPIC] == TEAM_LABEL_ENG_ABSENCE]
        if not eng_absence_rows.empty:
            abs_row = eng_absence_rows.iloc[0]
            has_any = any(
                col in abs_row.index and pd.notna(abs_row[col])
                for col in week_col_map.values()
            )
            if has_any:
                eng_absence_by_week = {
                    monday: (
                        float(abs_row[col])
                        if col in abs_row.index and pd.notna(abs_row[col])
                        else 0.0
                    )
                    for monday, col in week_col_map.items()
                }

    # --- engineer capacity resolution: week columns → scalar → headcount ---
    has_eng_bruto = TEAM_LABEL_ENGINEERS in config_df.index and pd.notna(config_df[TEAM_LABEL_ENGINEERS])
    has_num_eng   = TEAM_LABEL_NUM_ENGINEERS in config_df.index and pd.notna(config_df[TEAM_LABEL_NUM_ENGINEERS])

    if eng_bruto_by_week:
        num_engineers: float = sum(eng_bruto_by_week.values()) / len(eng_bruto_by_week)
        _log.info("Per-week bruto mode: scalar fallback set to %.2f PW (mean of Q%d weeks).",
                  num_engineers, quarter)
    elif has_eng_bruto:
        num_engineers = float(config_df[TEAM_LABEL_ENGINEERS])
    elif has_num_eng:
        num_engineers = float(config_df[TEAM_LABEL_NUM_ENGINEERS])
        _log.info("Deriving engineer capacity from '%s' = %s FTE.", TEAM_LABEL_NUM_ENGINEERS, num_engineers)
    else:
        raise ValueError(
            f"Input file missing engineer capacity: add a '{TEAM_LABEL_ENGINEERS}' row "
            f"with per-week values or a numeric Estimation, or use '{TEAM_LABEL_NUM_ENGINEERS}'."
        )

    # --- scalar management capacity (optional) ---
    if TEAM_LABEL_MANAGERS in config_df.index and pd.notna(config_df[TEAM_LABEL_MANAGERS]):
        num_managers: float = float(config_df[TEAM_LABEL_MANAGERS])
    else:
        num_managers = DEFAULT_MGMT_CAPACITY_PW
        _log.info("No '%s' row found — defaulting management capacity to %s PW/week.",
                  TEAM_LABEL_MANAGERS, DEFAULT_MGMT_CAPACITY_PW)

    # --- scalar absence (days → PW/week conversion) ---
    eng_absence_per_week: float | None = None
    if TEAM_LABEL_ENG_ABSENCE in config_df.index and pd.notna(config_df[TEAM_LABEL_ENG_ABSENCE]):
        eng_absence_per_week = float(config_df[TEAM_LABEL_ENG_ABSENCE]) / 5.0 / n_weeks

    mgmt_absence_per_week: float | None = None
    if TEAM_LABEL_MGMT_ABSENCE in config_df.index and pd.notna(config_df[TEAM_LABEL_MGMT_ABSENCE]):
        mgmt_absence_per_week = float(config_df[TEAM_LABEL_MGMT_ABSENCE]) / 5.0 / n_weeks

    capacity = CapacityConfig(
        num_engineers=num_engineers,
        num_managers=num_managers,
        eng_absence_per_week=eng_absence_per_week,
        mgmt_absence_per_week=mgmt_absence_per_week,
        eng_bruto_by_week=eng_bruto_by_week,
        eng_absence_by_week=eng_absence_by_week,
        q_weeks=frozenset(q_mondays),
    )

    # --- epic rows ---
    epics_df = df[~config_mask].reset_index(drop=True)
    epics_df = epics_df.dropna(how="all").reset_index(drop=True)

    no_description = epics_df[COL_EPIC].isna()
    for idx in epics_df[no_description].index:
        _log.warning("Discarding row %d: no '%s' value.", idx + 1, COL_EPIC)
    epics_df = epics_df[~no_description].reset_index(drop=True)

    # Rows without a Budget Bucket are annotations or computed rows, not epics.
    if COL_BUDGET_BUCKET in epics_df.columns:
        no_bucket = epics_df[COL_BUDGET_BUCKET].isna()
        for _, row in epics_df[no_bucket].iterrows():
            _log.warning(
                "Discarding row \"%s\": no '%s' value.", row[COL_EPIC], COL_BUDGET_BUCKET
            )
        epics_df = epics_df[~no_bucket].reset_index(drop=True)

    no_estimation = epics_df[COL_ESTIMATION].isna()
    for _, row in epics_df[no_estimation].iterrows():
        _log.warning(
            "Epic \"%s\": 'Estimation' is empty — defaulting to 0.", row[COL_EPIC]
        )
    epics_df[COL_ESTIMATION] = epics_df[COL_ESTIMATION].fillna(0.0)

    # --- Priority imputation ---
    # If the Priority column is absent entirely, add it; then fill any blank
    # cells from BUCKET_PRIORITY (keyed on Budget Bucket).  Unknown buckets
    # get 999 (lowest priority) with a warning.
    if COL_PRIORITY not in epics_df.columns:
        epics_df[COL_PRIORITY] = None
    if COL_BUDGET_BUCKET in epics_df.columns:
        missing_prio = epics_df[COL_PRIORITY].isna()
        for idx in epics_df[missing_prio].index:
            bucket = epics_df.at[idx, COL_BUDGET_BUCKET]
            imputed = BUCKET_PRIORITY.get(str(bucket), 999)
            if imputed == 999:
                _log.warning(
                    "Epic \"%s\": Priority blank and Budget Bucket %r not in BUCKET_PRIORITY "
                    "— defaulting to 999.",
                    epics_df.at[idx, COL_EPIC], bucket,
                )
            else:
                _log.info(
                    "Epic \"%s\": Priority blank — imputing %d from Budget Bucket %r.",
                    epics_df.at[idx, COL_EPIC], imputed, bucket,
                )
            epics_df.at[idx, COL_PRIORITY] = imputed

    missing = REQUIRED_INPUT_COLUMNS - set(epics_df.columns)
    if missing:
        raise ValueError(
            f"Input file is missing required columns: {sorted(missing)}"
        )

    return epics_df, capacity


def write_output(df: pd.DataFrame, path: Path) -> None:
    """Write the output allocation table to an Excel file (values only)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Allocation")
        ws = writer.sheets["Allocation"]
        _apply_conditional_formatting(
            ws,
            col_names=list(df.columns),
            labels=list(df[OUT_COL_EPIC]),
        )


def _apply_conditional_formatting(ws, col_names: list[str], labels: list[object]) -> None:
    """Apply formula-based formatting rules that remain dynamic after manual edits."""
    if not labels:
        return

    week_col_indices = [
        col_names.index(c) + 1
        for c in col_names if c not in _NON_WEEK_COLS
    ]
    if not week_col_indices:
        return

    epic_excel_rows = [
        i + 2 for i, lbl in enumerate(labels)
        if lbl not in _CAPACITY_LABELS and lbl is not None
    ]
    if not epic_excel_rows:
        return

    first_epic_row = epic_excel_rows[0]
    last_epic_row = epic_excel_rows[-1]
    first_data_row = 2
    last_data_row = len(labels) + 1
    first_col_letter = get_column_letter(1)
    last_col_letter = get_column_letter(len(col_names))
    full_data_range = f"{first_col_letter}{first_data_row}:{last_col_letter}{last_data_row}"

    first_week_letter = get_column_letter(week_col_indices[0])
    last_week_letter = get_column_letter(week_col_indices[-1])

    budget_bucket_col_letter = get_column_letter(col_names.index(OUT_COL_BUDGET_BUCKET) + 1)
    off_estimate_col_letter = get_column_letter(col_names.index(OUT_COL_OFF_ESTIMATE) + 1)

    r_capacity_alert = labels.index(LABEL_CAPACITY_ALERT_ROW) + 2

    # Highlight boolean outputs when they evaluate to TRUE.
    off_estimate_range = (
        f"{off_estimate_col_letter}{first_epic_row}:"
        f"{off_estimate_col_letter}{last_epic_row}"
    )
    off_estimate_true_formula = f"{off_estimate_col_letter}{first_epic_row}=TRUE"
    ws.conditional_formatting.add(
        off_estimate_range,
        FormulaRule(
            formula=[off_estimate_true_formula],
            stopIfTrue=True,
            fill=_ALERT_FILL,
            font=_ALERT_FONT,
        ),
    )

    capacity_alert_range = (
        f"{first_week_letter}{r_capacity_alert}:"
        f"{last_week_letter}{r_capacity_alert}"
    )
    capacity_alert_true_formula = f"{first_week_letter}{r_capacity_alert}=TRUE"
    ws.conditional_formatting.add(
        capacity_alert_range,
        FormulaRule(
            formula=[capacity_alert_true_formula],
            stopIfTrue=True,
            fill=_ALERT_FILL,
            font=_ALERT_FONT,
        ),
    )

    # Color full data rows by Budget Bucket text; formulas are relative by row.
    for bucket_label, fill_color in _BUCKET_COLOR_BY_LABEL:
        safe_label = bucket_label.replace('"', '""')
        ws.conditional_formatting.add(
            full_data_range,
            FormulaRule(
                formula=[f'${budget_bucket_col_letter}{first_data_row}="{safe_label}"'],
                stopIfTrue=False,
                fill=PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color),
            ),
        )


def write_output_with_formulas(df: pd.DataFrame, path: Path, n_base_weeks: int) -> None:
    """
    Write the output allocation table to an Excel file with formulas for
    calculated fields:

    - Engineer Net Capacity row: ``=<bruto_cell>-<absence_cell>`` per week
    - Management Net Capacity row: same pattern
    - Total Weeks column (capacity + epic rows): ``=SUM(<first_week>:<last_Q_week>)``
      where ``last_Q_week`` is the last column of the requested quarter only
      (overflow columns are excluded from the Total Weeks sum).
    - Weekly Allocation row: ``=SUM(<first_epic>:<last_epic>)`` per week
    """
    # Write values first, then reopen to replace calculated cells with formulas.
    write_output(df, path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    col_names = list(df.columns)
    week_col_indices = [
        col_names.index(c) + 1
        for c in col_names if c not in _NON_WEEK_COLS
    ]
    total_weeks_col_idx = col_names.index(OUT_COL_TOTAL_WEEKS) + 1

    # Excel row = DataFrame index + 2  (0-indexed df + 1 for header + 1 for 1-based)
    labels = list(df[OUT_COL_EPIC])

    def excel_row(label: str) -> int:
        return labels.index(label) + 2

    r_eng_bruto   = excel_row(LABEL_ENG_BRUTO)
    r_eng_absence = excel_row(LABEL_ENG_ABSENCE)
    r_eng_net     = excel_row(LABEL_ENG_NET)
    r_mgmt_cap    = excel_row(LABEL_MGMT_CAPACITY)
    r_mgmt_absence = excel_row(LABEL_MGMT_ABSENCE)
    r_mgmt_net    = excel_row(LABEL_MGMT_NET)
    r_total       = excel_row(LABEL_TOTAL_ROW)

    epic_excel_rows = [
        i + 2 for i, lbl in enumerate(labels) if lbl not in _CAPACITY_LABELS
    ]
    first_week_letter = get_column_letter(week_col_indices[0])
    last_q_week_letter = get_column_letter(week_col_indices[n_base_weeks - 1])
    first_epic_row = epic_excel_rows[0]
    last_epic_row  = epic_excel_rows[-1]

    # Engineer Net Capacity: =<bruto> - <absence>
    for ci in week_col_indices:
        cl = get_column_letter(ci)
        ws.cell(r_eng_net, ci).value = f"={cl}{r_eng_bruto}-{cl}{r_eng_absence}"

    # Management Net Capacity: =<mgmt_cap> - <mgmt_absence>
    for ci in week_col_indices:
        cl = get_column_letter(ci)
        ws.cell(r_mgmt_net, ci).value = f"={cl}{r_mgmt_cap}-{cl}{r_mgmt_absence}"

    # Total Weeks for each epic: =SUM(<first_week_col><row>:<last_Q_week_col><row>)
    for er in epic_excel_rows:
        ws.cell(er, total_weeks_col_idx).value = (
            f"=SUM({first_week_letter}{er}:{last_q_week_letter}{er})"
        )

    # Total Weeks for capacity header rows: same SUM — Q-only
    for r_cap in (r_eng_bruto, r_eng_absence, r_eng_net, r_mgmt_cap, r_mgmt_absence, r_mgmt_net):
        ws.cell(r_cap, total_weeks_col_idx).value = (
            f"=SUM({first_week_letter}{r_cap}:{last_q_week_letter}{r_cap})"
        )

    # Weekly Allocation row: =SUM(<col><first_epic>:<col><last_epic>)
    for ci in week_col_indices:
        cl = get_column_letter(ci)
        ws.cell(r_total, ci).value = (
            f"=SUM({cl}{first_epic_row}:{cl}{last_epic_row})"
        )

    # Total row Estimation and Total Weeks: sum over all epic rows
    estimation_col_idx = col_names.index(OUT_COL_ESTIMATION) + 1
    estimation_col_letter = get_column_letter(estimation_col_idx)
    total_weeks_col_letter = get_column_letter(total_weeks_col_idx)
    ws.cell(r_total, estimation_col_idx).value = (
        f"=SUM({estimation_col_letter}{first_epic_row}:{estimation_col_letter}{last_epic_row})"
    )
    ws.cell(r_total, total_weeks_col_idx).value = (
        f"=SUM({total_weeks_col_letter}{first_epic_row}:{total_weeks_col_letter}{last_epic_row})"
    )

    # Off Estimate column (epic rows only): =ABS(<total_weeks_cell>-<estimation_cell>)>0.05
    off_estimate_col_idx = col_names.index(OUT_COL_OFF_ESTIMATE) + 1
    for er in epic_excel_rows:
        ws.cell(er, off_estimate_col_idx).value = (
            f"=ABS({total_weeks_col_letter}{er}-{estimation_col_letter}{er})>0.05"
        )

    # Off Capacity row (week columns): =ABS(<weekly_alloc_cell>-<eng_net_cell>)>0.1
    r_capacity_alert = excel_row(LABEL_CAPACITY_ALERT_ROW)
    for ci in week_col_indices:
        cl = get_column_letter(ci)
        ws.cell(r_capacity_alert, ci).value = (
            f"=ABS({cl}{r_total}-{cl}{r_eng_net})>0.1"
        )

    wb.save(path)
